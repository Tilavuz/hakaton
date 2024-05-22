import io
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils.cell import get_column_letter
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from accounts.permissions import (
    IsAdmin,
    IsHokim,
    IsHokimYordamchisi,
    IsTumanMasul,
    IsTumanYoshlarIshlari,
    IsMahallaMasul,
    IsMaktabMasul,
    IsMahallaYoshlarIshlari,
    IsMaktabYoshlarIshlari,
)
from django.http import FileResponse
from adduser.models import UserInfo, Certificate
from accounts.models import Tuman, Mahalla, Maktab, User


# View to export Tuman data to Excel
class TumanExcelExportView(APIView):
    # Set the required permissions for the view
    permission_classes = [
        IsAuthenticated,
        IsAdmin | IsHokim | IsHokimYordamchisi | IsTumanMasul | IsTumanYoshlarIshlari,
    ]

    def get(self, request):
        # Fetch all Tuman records
        tumans = Tuman.objects.all()
        workbook = openpyxl.Workbook()
        tuman_worksheet = workbook.active
        tuman_worksheet.title = "Tuman Data"
        # Add data to the worksheet
        self.add_tuman_data(tuman_worksheet, tumans)
        # Save workbook to a BytesIO stream
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        # Return the file as an attachment in the response
        response = FileResponse(
            excel_file, as_attachment=True, filename="tuman_data.xlsx"
        )
        return response

    def add_tuman_data(self, worksheet, tumans):
        # Define header row
        headers = ["Tuman Nomi", "Viloyat", "Umumiy reja", "Ingliz tili B2 rejasi", 
                   "Ingliz tili C1 rejasi", "Ingliz tili C2 rejasi", 
                   "Nemis tili va boshqa tillar B2 reja", "Nemis tili va boshqa tillar C1 reja", 
                   "Nemis tili va boshqa tillar C2 reja"]
        # Add headers to the worksheet
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

        # Add Tuman data to the worksheet
        row = 2
        for tuman in tumans:
            worksheet.cell(row=row, column=1, value=tuman.name)
            worksheet.cell(row=row, column=2, value="Qashqadaryo")
            worksheet.cell(row=row, column=3, value=tuman.overall)
            worksheet.cell(row=row, column=4, value=tuman.plan_en_b2)
            worksheet.cell(row=row, column=5, value=tuman.plan_en_c1)
            worksheet.cell(row=row, column=6, value=tuman.plan_en_c2)
            worksheet.cell(row=row, column=7, value=tuman.plan_deorother_b2)
            worksheet.cell(row=row, column=8, value=tuman.plan_deorother_c1)
            worksheet.cell(row=row, column=9, value=tuman.plan_deorother_c2)
            row += 1
        
        # Adjust column widths
        for col in range(1, 10):
            column_letter = get_column_letter(col)
            worksheet.column_dimensions[column_letter].width = 20


# View to export Mahalla data to Excel
class MahallaExcelExportView(APIView):
    # Set the required permissions for the view
    permission_classes = [
        IsAuthenticated,
        IsAdmin
        | IsHokim
        | IsHokimYordamchisi
        | IsTumanMasul
        | IsTumanYoshlarIshlari
        | IsMahallaMasul
        | IsMahallaYoshlarIshlari,
    ]

    def get(self, request):
        # Fetch all Mahalla records
        mahallas = Mahalla.objects.all()
        workbook = openpyxl.Workbook()
        mahalla_worksheet = workbook.active
        mahalla_worksheet.title = "Mahalla Data"
        # Add data to the worksheet
        self.add_mahalla_data(mahalla_worksheet, mahallas)
        # Save workbook to a BytesIO stream
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        # Return the file as an attachment in the response
        response = FileResponse(
            excel_file, as_attachment=True, filename="mahalla_data.xlsx"
        )
        return response

    def add_mahalla_data(self, worksheet, mahallas):
        # Define header row
        headers = ["Mahalla Nomi", "Tuman", "Umumiy reja", "Ingliz tili B2 rejasi", 
                   "Ingliz tili C1 rejasi", "Ingliz tili C2 rejasi", 
                   "Nemis tili va boshqa tillar B2 reja", "Nemis tili va boshqa tillar C1 reja", 
                   "Nemis tili va boshqa tillar C2 reja"]
        # Add headers to the worksheet
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

        # Add Mahalla data to the worksheet
        row = 2
        for mahalla in mahallas:
            worksheet.cell(row=row, column=1, value=mahalla.name)
            worksheet.cell(row=row, column=2, value=mahalla.tuman.name)
            worksheet.cell(row=row, column=3, value=mahalla.overall)
            worksheet.cell(row=row, column=4, value=mahalla.plan_en_b2)
            worksheet.cell(row=row, column=5, value=mahalla.plan_en_c1)
            worksheet.cell(row=row, column=6, value=mahalla.plan_en_c2)
            worksheet.cell(row=row, column=7, value=mahalla.plan_deorother_b2)
            worksheet.cell(row=row, column=8, value=mahalla.plan_deorother_c1)
            worksheet.cell(row=row, column=9, value=mahalla.plan_deorother_c2)
            row += 1
        
        # Adjust column widths
        for col in range(1, 10):
            column_letter = get_column_letter(col)
            worksheet.column_dimensions[column_letter].width = 20


# View to export Maktab data to Excel
class MaktabExcelExportView(APIView):
    # Set the required permissions for the view
    permission_classes = [
        IsAuthenticated,
        IsAdmin
        | IsHokim
        | IsHokimYordamchisi
        | IsTumanMasul
        | IsTumanYoshlarIshlari
        | IsMahallaMasul
        | IsMahallaYoshlarIshlari
        | IsMaktabMasul
        | IsMaktabYoshlarIshlari,
    ]

    def get(self, request):
        # Fetch all Maktab records
        maktabs = Maktab.objects.all()
        workbook = openpyxl.Workbook()
        maktab_worksheet = workbook.active
        maktab_worksheet.title = "Maktab Data"
        # Add data to the worksheet
        self.add_maktab_data(maktab_worksheet, maktabs)
        # Save workbook to a BytesIO stream
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        # Return the file as an attachment in the response
        response = FileResponse(
            excel_file, as_attachment=True, filename="maktab_data.xlsx"
        )
        return response

    def add_maktab_data(self, worksheet, maktabs):
        # Define header row
        headers = ["Maktab Nomi", "Mahalla", "Umumiy reja", "Ingliz tili B2 rejasi", 
                   "Ingliz tili C1 rejasi", "Ingliz tili C2 rejasi", 
                   "Nemis tili va boshqa tillar B2 reja", "Nemis tili va boshqa tillar C1 reja", 
                   "Nemis tili va boshqa tillar C2 reja"]
        # Add headers to the worksheet
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

        # Add Maktab data to the worksheet
        row = 2
        for maktab in maktabs:
            worksheet.cell(row=row, column=1, value=maktab.name)
            worksheet.cell(row=row, column=2, value=maktab.mahalla.name)
            worksheet.cell(row=row, column=3, value=maktab.overall)
            worksheet.cell(row=row, column=4, value=maktab.plan_en_b2)
            worksheet.cell(row=row, column=5, value=maktab.plan_en_c1)
            worksheet.cell(row=row, column=6, value=maktab.plan_en_c2)
            worksheet.cell(row=row, column=7, value=maktab.plan_deorother_b2)
            worksheet.cell(row=row, column=8, value=maktab.plan_deorother_c1)
            worksheet.cell(row=row, column=9, value=maktab.plan_deorother_c2)
            row += 1
        
        # Adjust column widths
        for col in range(1, 10):
            column_letter = get_column_letter(col)
            worksheet.column_dimensions[column_letter].width = 20


# View to export user data to Excel
class UserExcelExportView(APIView):
    # Set the required permissions for the view
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Fetch all User records
        users = User.objects.all()
        workbook = openpyxl.Workbook()
        user_worksheet = workbook.active
        user_worksheet.title = "User Data"
        # Add data to the worksheet
        self.add_user_data(user_worksheet, users)
        # Save workbook to a BytesIO stream
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        # Return the file as an attachment in the response
        response = FileResponse(
            excel_file, as_attachment=True, filename="user_data.xlsx"
        )
        return response

    def add_user_data(self, worksheet, users):
        # Define header row
        headers = ["F.I.O", "Tuman", "Mahalla", "Maktab", "Passport", "Viloyat", 
                   "Yunalish", "Jinsi", "Tug'ilgan yil", "Telefon raqam", "Sertifikat nomi", 
                   "Muassasa", "Daraja", "Seriya", "Kelgan vaqti"]
        # Add headers to the worksheet
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

        # Add User data to the worksheet
        row = 2
        for user in users:
            userinfo = user.userinfo
            certificates = Certificate.objects.filter(user=user)
            certificates_str = ", ".join([cert.certificate_name for cert in certificates])
            worksheet.cell(row=row, column=1, value=f"{userinfo.first_name} {userinfo.last_name}")
            worksheet.cell(row=row, column=2, value=userinfo.tuman.name)
            worksheet.cell(row=row, column=3, value=userinfo.mahalla.name)
            worksheet.cell(row=row, column=4, value=userinfo.maktab.name)
            worksheet.cell(row=row, column=5, value=userinfo.passport)
            worksheet.cell(row=row, column=6, value="Qashqadaryo")
            worksheet.cell(row=row, column=7, value=userinfo.yunalish)
            worksheet.cell(row=row, column=8, value=userinfo.jinsi)
            worksheet.cell(row=row, column=9, value=userinfo.birth_date.strftime("%Y-%m-%d"))
            worksheet.cell(row=row, column=10, value=userinfo.phone_number)
            worksheet.cell(row=row, column=11, value=certificates_str)
            worksheet.cell(row=row, column=12, value=userinfo.muassasa)
            worksheet.cell(row=row, column=13, value=userinfo.daraja)
            worksheet.cell(row=row, column=14, value=userinfo.seriya)
            worksheet.cell(row=row, column=15, value=userinfo.created_at.strftime("%Y-%m-%d %H:%M:%S"))
            row += 1
        
        # Adjust column widths
        for col in range(1, 16):
            column_letter = get_column_letter(col)
            worksheet.column_dimensions[column_letter].width = 20
