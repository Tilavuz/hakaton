import io
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils.cell import get_column_letter
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from accounts.permissions import (
    IsAdmin,
    IsHokim,
    IsHokimYordamchisi,
    IsTumanMasul,
    IsTumanYoshlarIshlari,
    IsMahallaMasul,
    IsMaktabMasul,
    IsMahallaYoshlarIshlari,
    IsMaktabYoshlarIshlari,
)
from django.http import FileResponse
from adduser.models import UserInfo, Certificate
from accounts.models import Tuman, Mahalla, Maktab, User


class TumanExcelExportView(APIView):
    permission_classes = [
        IsAuthenticated,
        IsAdmin | IsHokim | IsHokimYordamchisi | IsTumanMasul | IsTumanYoshlarIshlari,
    ]

    def get(self, request):
        tumans = Tuman.objects.all()
        workbook = openpyxl.Workbook()
        tuman_worksheet = workbook.active
        tuman_worksheet.title = "Tuman Data"
        self.add_tuman_data(tuman_worksheet, tumans)
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        response = FileResponse(
            excel_file, as_attachment=True, filename="tuman_data.xlsx"
        )
        return response

    def add_tuman_data(self, worksheet, tumans):
        worksheet["A1"] = "Tuman Nomi"
        worksheet["B1"] = "Viloyat"
        worksheet["C1"] = "Umumiy reja"
        worksheet["D1"] = "Ingliz tili B2 rejasi"
        worksheet["E1"] = "Ingliz tili C1 rejasi"
        worksheet["F1"] = "Ingliz tili C2 rejasi"
        worksheet["G1"] = "Nemis tili va boshqa tillar B2 reja"
        worksheet["H1"] = "Nemis tili va boshqa tillar C1 reja"
        worksheet["I1"] = "Nemis tili va boshqa tillar C2 reja"
        worksheet["A1"].font = Font(bold=True)
        worksheet["B1"].font = Font(bold=True)
        worksheet["C1"].font = Font(bold=True)
        worksheet["D1"].font = Font(bold=True)
        worksheet["E1"].font = Font(bold=True)
        worksheet["F1"].font = Font(bold=True)
        worksheet["G1"].font = Font(bold=True)
        worksheet["H1"].font = Font(bold=True)
        row = 2
        for tuman in tumans:
            worksheet.cell(row=row, column=1, value=tuman.name)
            worksheet.cell(row=row, column=2, value="Qashqadaryo")
            worksheet.cell(row=row, column=3, value=tuman.overall)
            worksheet.cell(row=row, column=4, value=tuman.plan_en_b2)
            worksheet.cell(row=row, column=5, value=tuman.plan_en_c1)
            worksheet.cell(row=row, column=6, value=tuman.plan_en_c2)
            worksheet.cell(row=row, column=7, value=tuman.plan_deorother_b2)
            worksheet.cell(row=row, column=8, value=tuman.plan_deorother_c1)
            worksheet.cell(row=row, column=9, value=tuman.plan_deorother_c2)
            row += 1
        for col in range(1, 9):
            column_letter = get_column_letter(col)
            worksheet.column_dimensions[column_letter].width = 20


class MahallaExcelExportView(APIView):
    permission_classes = [
        IsAuthenticated,
        IsAdmin
        | IsHokim
        | IsHokimYordamchisi
        | IsTumanMasul
        | IsTumanYoshlarIshlari
        | IsMahallaMasul
        | IsMahallaYoshlarIshlari,
    ]

    def get(self, request):
        mahallas = Mahalla.objects.all()
        workbook = openpyxl.Workbook()
        mahalla_worksheet = workbook.active
        mahalla_worksheet.title = "Mahalla Data"
        self.add_mahalla_data(mahalla_worksheet, mahallas)
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        response = FileResponse(
            excel_file, as_attachment=True, filename="mahalla_data.xlsx"
        )
        return response

    def add_mahalla_data(self, worksheet, mahallas):
        worksheet["A1"] = "Mahalla Nomi"
        worksheet["B1"] = "Tuman"
        worksheet["C1"] = "Umumiy reja"
        worksheet["D1"] = "Ingliz tili B2 rejasi"
        worksheet["E1"] = "Ingliz tili C1 rejasi"
        worksheet["F1"] = "Ingliz tili C2 rejasi"
        worksheet["G1"] = "Nemis tili va boshqa tillar B2 reja"
        worksheet["H1"] = "Nemis tili va boshqa tillar C1 reja"
        worksheet["I1"] = "Nemis tili va boshqa tillar C2 reja"
        worksheet["A1"].font = Font(bold=True)
        worksheet["B1"].font = Font(bold=True)
        worksheet["C1"].font = Font(bold=True)
        worksheet["D1"].font = Font(bold=True)
        worksheet["E1"].font = Font(bold=True)
        worksheet["F1"].font = Font(bold=True)
        worksheet["G1"].font = Font(bold=True)
        worksheet["H1"].font = Font(bold=True)
        row = 2
        for mahalla in mahallas:
            worksheet.cell(row=row, column=1, value=mahalla.name)
            worksheet.cell(row=row, column=2, value=mahalla.tuman.name)
            worksheet.cell(row=row, column=3, value=mahalla.overall)
            worksheet.cell(row=row, column=4, value=mahalla.plan_en_b2)
            worksheet.cell(row=row, column=5, value=mahalla.plan_en_c1)
            worksheet.cell(row=row, column=6, value=mahalla.plan_en_c2)
            worksheet.cell(row=row, column=7, value=mahalla.plan_deorother_b2)
            worksheet.cell(row=row, column=8, value=mahalla.plan_deorother_c1)
            worksheet.cell(row=row, column=9, value=mahalla.plan_deorother_c2)
            row += 1
        for col in range(1, 10):
            column_letter = get_column_letter(col)
            worksheet.column_dimensions[column_letter].width = 20


class MaktabExcelExportView(APIView):
    permission_classes = [
        IsAuthenticated,
        IsAdmin
        | IsHokim
        | IsHokimYordamchisi
        | IsTumanMasul
        | IsTumanYoshlarIshlari
        | IsMahallaMasul
        | IsMahallaYoshlarIshlari
        | IsMaktabMasul
        | IsMaktabYoshlarIshlari,
    ]

    def get(self, request):
        maktabs = Maktab.objects.all()
        workbook = openpyxl.Workbook()
        maktab_worksheet = workbook.active
        maktab_worksheet.title = "Maktab Data"
        self.add_maktab_data(maktab_worksheet, maktabs)
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        response = FileResponse(
            excel_file, as_attachment=True, filename="maktab_data.xlsx"
        )
        return response

    def add_maktab_data(self, worksheet, maktabs):
        worksheet["A1"] = "Maktab Nomi"
        worksheet["B1"] = "Mahalla"
        worksheet["C1"] = "Umumiy reja"
        worksheet["D1"] = "Ingliz tili B2 rejasi"
        worksheet["E1"] = "Ingliz tili C1 rejasi"
        worksheet["F1"] = "Ingliz tili C2 rejasi"
        worksheet["G1"] = "Nemis tili va boshqa tillar B2 reja"
        worksheet["H1"] = "Nemis tili va boshqa tillar C1 reja"
        worksheet["I1"] = "Nemis tili va boshqa tillar C2 reja"
        worksheet["A1"].font = Font(bold=True)
        worksheet["B1"].font = Font(bold=True)
        worksheet["C1"].font = Font(bold=True)
        worksheet["D1"].font = Font(bold=True)
        worksheet["E1"].font = Font(bold=True)
        worksheet["F1"].font = Font(bold=True)
        worksheet["G1"].font = Font(bold=True)
        worksheet["H1"].font = Font(bold=True)
        row = 2
        for maktab in maktabs:
            worksheet.cell(row=row, column=1, value=maktab.name)
            worksheet.cell(row=row, column=2, value=maktab.mahalla.name)
            worksheet.cell(row=row, column=3, value=maktab.overall)
            worksheet.cell(row=row, column=4, value=maktab.plan_en_b2)
            worksheet.cell(row=row, column=5, value=maktab.plan_en_c1)
            worksheet.cell(row=row, column=6, value=maktab.plan_en_c2)
            worksheet.cell(row=row, column=7, value=maktab.plan_deorother_b2)
            worksheet.cell(row=row, column=8, value=maktab.plan_deorother_c1)
            worksheet.cell(row=row, column=9, value=maktab.plan_deorother_c2)
            row += 1
        for col in range(1, 10):
            column_letter = get_column_letter(col)
            worksheet.column_dimensions[column_letter].width = 20


class UserExcelExportView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin | IsHokim | IsHokimYordamchisi]

    def get(self, request):
        users = User.objects.all()
        workbook = openpyxl.Workbook()
        user_worksheet = workbook.active
        user_worksheet.title = "User Data"
        self.add_user_data(user_worksheet, users)
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        response = FileResponse(
            excel_file, as_attachment=True, filename="user_data.xlsx"
        )
        return response

    def add_user_data(self, worksheet, users):
        worksheet["A1"] = "Username"
        worksheet["B1"] = "Ismi"
        worksheet["C1"] = "Familyasi"
        worksheet["D1"] = "Oxirgi login"
        worksheet["E1"] = "Tuman"
        worksheet["F1"] = "Role"
        worksheet["A1"].font = Font(bold=True)
        worksheet["B1"].font = Font(bold=True)
        worksheet["C1"].font = Font(bold=True)
        worksheet["D1"].font = Font(bold=True)
        worksheet["E1"].font = Font(bold=True)
        worksheet["F1"].font = Font(bold=True)
        row = 2
        for user in users:
            worksheet.cell(row=row, column=1, value=user.username)
            worksheet.cell(row=row, column=2, value=user.first_name)
            worksheet.cell(row=row, column=3, value=user.last_name)
            last_login = user.last_login.replace(tzinfo=None)
            worksheet.cell(row=row, column=4, value=last_login)
            worksheet.cell(
                row=row, column=5, value=user.tuman.name if user.tuman else ""
            )
            worksheet.cell(row=row, column=6, value=dict(user.ROLE_CHOICES)[user.role])
            row += 1
        for col in range(1, 7):
            column_letter = get_column_letter(col)
            worksheet.column_dimensions[column_letter].width = 20


class StatisticsExcelExportView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin | IsHokim | IsHokimYordamchisi]

    def get(self, request):
        workbook = openpyxl.Workbook()
        statistics_worksheet = workbook.active
        statistics_worksheet.title = "Statistics"
        self.add_statistics_data(statistics_worksheet)
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        response = FileResponse(
            excel_file, as_attachment=True, filename="statistics.xlsx"
        )
        return response

    def add_statistics_data(self, worksheet):
        worksheet["A1"] = "Tuman"
        worksheet["B1"] = "Mahalla"
        worksheet["C1"] = "Maktab"
        worksheet["D1"] = "Ingliz tili B2 rejasi"
        worksheet["E1"] = "Ingliz tili C1 rejasi"
        worksheet["F1"] = "Ingliz tili C2 rejasi"
        worksheet["G1"] = "Nemis tili va boshqa tillar B2 reja"
        worksheet["H1"] = "Nemis tili va boshqa tillar C1 reja"
        worksheet["I1"] = "Nemis tili va boshqa tillar C2 reja"
        worksheet["A1"].font = Font(bold=True)
        worksheet["B1"].font = Font(bold=True)
        worksheet["C1"].font = Font(bold=True)
        worksheet["D1"].font = Font(bold=True)
        worksheet["E1"].font = Font(bold=True)
        worksheet["F1"].font = Font(bold=True)
        worksheet["G1"].font = Font(bold=True)
        worksheet["H1"].font = Font(bold=True)
        row = 2
        tumans = Tuman.objects.all()
        for tuman in tumans:
            worksheet.cell(row=row, column=1, value=tuman.name)
            worksheet.cell(row=row, column=2, value=tuman.mahalla_set.count())
            worksheet.cell(row=row, column=3, value=tuman.maktab_set.count())
            worksheet.cell(row=row, column=4, value=tuman.plan_en_b2)
            worksheet.cell(row=row, column=5, value=tuman.plan_en_c1)
            worksheet.cell(row=row, column=6, value=tuman.plan_en_c2)
            worksheet.cell(row=row, column=7, value=tuman.plan_deorother_b2)
            worksheet.cell(row=row, column=8, value=tuman.plan_deorother_c1)
            worksheet.cell(row=row, column=9, value=tuman.plan_deorother_c2)
            row += 1
        for col in range(1, 10):
            column_letter = get_column_letter(col)
            worksheet.column_dimensions[column_letter].width = 20


class CertificateExcelExportView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin | IsHokim | IsHokimYordamchisi]

    def get(self, request):
        workbook = openpyxl.Workbook()
        certificate_worksheet = workbook.active
        certificate_worksheet.title = "Certificates"
        self.add_certificate_data(certificate_worksheet)
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        response = FileResponse(
            excel_file, as_attachment=True, filename="certificates.xlsx"
        )
        return response

    def add_certificate_data(self, worksheet):
        worksheet["A1"] = "Tuman"
        worksheet["B1"] = "Ingliz tili B2 rejasi"
        worksheet["C1"] = "Ingliz tili C1 rejasi"
        worksheet["D1"] = "Ingliz tili C2 rejasi"
        worksheet["E1"] = "Nemis tili va boshqa tillar B2 reja"
        worksheet["F1"] = "Nemis tili va boshqa tillar C1 reja"
        worksheet["G1"] = "Nemis tili va boshqa tillar C2 reja"
        worksheet["A1"].font = Font(bold=True)
        worksheet["B1"].font = Font(bold=True)
        worksheet["C1"].font = Font(bold=True)
        worksheet["D1"].font = Font(bold=True)
        worksheet["E1"].font = Font(bold=True)
        worksheet["F1"].font = Font(bold=True)
        worksheet["G1"].font = Font(bold=True)
        row = 2
        tumans = Tuman.objects.all()
        for tuman in tumans:
            certificates = Certificate.objects.filter(user_info__status="uqimoqda")
            worksheet.cell(row=row, column=1, value=tuman.name)
            worksheet.cell(
                row=row,
                column=2,
                value=certificates.filter(title="english", overel="B2").count(),
            )
            worksheet.cell(
                row=row,
                column=3,
                value=certificates.filter(title="english", overel="C1").count(),
            )
            worksheet.cell(
                row=row,
                column=4,
                value=certificates.filter(title="english", overel="C2").count(),
            )
            worksheet.cell(
                row=row,
                column=5,
                value=certificates.filter(title="nemesis", overel="B2").count(),
            )
            worksheet.cell(
                row=row,
                column=6,
                value=certificates.filter(title="nemesis", overel="C1").count(),
            )
            worksheet.cell(
                row=row,
                column=7,
                value=certificates.filter(title="nemesis", overel="C2").count(),
            )
            row += 1
        for col in range(1, 8):
            column_letter = get_column_letter(col)
            worksheet.column_dimensions[column_letter].width = 20


class UserInfoExcelExportView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin | IsHokim | IsHokimYordamchisi]

    def get(self, request):
        workbook = openpyxl.Workbook()
        user_info_worksheet = workbook.active
        user_info_worksheet.title = "User Info"
        self.add_user_info_data(user_info_worksheet)
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        response = FileResponse(
            excel_file, as_attachment=True, filename="user_info.xlsx"
        )
        return response

    def add_user_info_data(self, worksheet):
        worksheet["A1"] = "Tuman"
        worksheet["B1"] = "English Uqimoqda"
        worksheet["C1"] = "English Tugatgan"
        worksheet["D1"] = "English A1"
        worksheet["E1"] = "English B2"
        worksheet["F1"] = "Nemesis Uqimoqda"
        worksheet["G1"] = "Nemesis Tugatgan"
        worksheet["H1"] = "Nemesis A1"
        worksheet["I1"] = "Nemesis B2"
        worksheet["J1"] = "Others Uqimoqda"
        worksheet["K1"] = "Others Tugatgan"
        worksheet["L1"] = "Others A1"
        worksheet["M1"] = "Others B2"
        worksheet["A1"].font = Font(bold=True)
        worksheet["B1"].font = Font(bold=True)
        worksheet["C1"].font = Font(bold=True)
        worksheet["D1"].font = Font(bold=True)
        worksheet["E1"].font = Font(bold=True)
        worksheet["F1"].font = Font(bold=True)
        worksheet["G1"].font = Font(bold=True)
        worksheet["H1"].font = Font(bold=True)
        worksheet["I1"].font = Font(bold=True)
        worksheet["J1"].font = Font(bold=True)
        worksheet["K1"].font = Font(bold=True)
        worksheet["L1"].font = Font(bold=True)
        worksheet["M1"].font = Font(bold=True)
        row = 2
        tumans = Tuman.objects.all()
        for tuman in tumans:
            worksheet.cell(row=row, column=1, value=tuman.name)
            worksheet.cell(
                row=row,
                column=2,
                value=UserInfo.objects.filter(
                    title="english", status="uqimoqda", tuman=tuman
                ).count(),
            )
            worksheet.cell(
                row=row,
                column=3,
                value=UserInfo.objects.filter(
                    title="english", status="tugatgan", tuman=tuman
                ).count(),
            )
            worksheet.cell(
                row=row,
                column=4,
                value=Certificate.objects.filter(
                    user_info__title="english", overel="A1", user_info__tuman=tuman
                ).count(),
            )
            worksheet.cell(
                row=row,
                column=5,
                value=Certificate.objects.filter(
                    user_info__title="english", overel="B2", user_info__tuman=tuman
                ).count(),
            )
            worksheet.cell(
                row=row,
                column=6,
                value=UserInfo.objects.filter(
                    title="nemesis", status="uqimoqda", tuman=tuman
                ).count(),
            )
            worksheet.cell(
                row=row,
                column=7,
                value=UserInfo.objects.filter(
                    title="nemesis", status="tugatgan", tuman=tuman
                ).count(),
            )
            worksheet.cell(
                row=row,
                column=8,
                value=Certificate.objects.filter(
                    user_info__title="nemesis", overel="A1", user_info__tuman=tuman
                ).count(),
            )
            worksheet.cell(
                row=row,
                column=9,
                value=Certificate.objects.filter(
                    user_info__title="nemesis", overel="B2", user_info__tuman=tuman
                ).count(),
            )
            worksheet.cell(
                row=row,
                column=10,
                value=UserInfo.objects.filter(
                    title="others", status="uqimoqda", tuman=tuman
                ).count(),
            )
            worksheet.cell(
                row=row,
                column=11,
                value=UserInfo.objects.filter(
                    title="others", status="tugatgan", tuman=tuman
                ).count(),
            )
            worksheet.cell(
                row=row,
                column=12,
                value=Certificate.objects.filter(
                    user_info__title="others", overel="A1", user_info__tuman=tuman
                ).count(),
            )
            worksheet.cell(
                row=row,
                column=13,
                value=Certificate.objects.filter(
                    user_info__title="others", overel="B2", user_info__tuman=tuman
                ).count(),
            )
            row += 1
        for col in range(1, 14):
            column_letter = get_column_letter(col)
            worksheet.column_dimensions[column_letter].width = 20


class NemesisExcelExportView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin | IsHokim | IsHokimYordamchisi]

    def get(self, request):
        workbook = openpyxl.Workbook()
        nemesis_and_others_worksheet = workbook.active
        nemesis_and_others_worksheet.title = "Nemesis and Others"
        self.add_nemesis_and_others_data(nemesis_and_others_worksheet)
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        response = FileResponse(
            excel_file, as_attachment=True, filename="nemesis_and_others.xlsx"
        )
        return response

    def add_nemesis_and_others_data(self, worksheet):
        worksheet["A1"] = "Tuman"
        worksheet["B1"] = "Nemesis and Others Uqimoqda"
        worksheet["C1"] = "Nemesis and Others Tugatgan"
        worksheet["D1"] = "Nemesis and Others A1"
        worksheet["E1"] = "Nemesis and Others B2"
        worksheet["A1"].font = Font(bold=True)
        worksheet["B1"].font = Font(bold=True)
        worksheet["C1"].font = Font(bold=True)
        worksheet["D1"].font = Font(bold=True)
        worksheet["E1"].font = Font(bold=True)
        row = 2
        tumans = Tuman.objects.all()
        for tuman in tumans:
            worksheet.cell(row=row, column=1, value=tuman.name)
            worksheet.cell(
                row=row,
                column=2,
                value=UserInfo.objects.filter(
                    title__in=["nemesis", "others"], status="uqimoqda", tuman=tuman
                ).count(),
            )
            worksheet.cell(
                row=row,
                column=3,
                value=UserInfo.objects.filter(
                    title__in=["nemesis", "others"], status="tugatgan", tuman=tuman
                ).count(),
            )
            worksheet.cell(
                row=row,
                column=4,
                value=Certificate.objects.filter(
                    user_info__title__in=["nemesis", "others"],
                    overel="A1",
                    user_info__tuman=tuman,
                ).count(),
            )
            worksheet.cell(
                row=row,
                column=5,
                value=Certificate.objects.filter(
                    user_info__title__in=["nemesis", "others"],
                    overel="B2",
                    user_info__tuman=tuman,
                ).count(),
            )
            row += 1
        for col in range(1, 6):
            column_letter = get_column_letter(col)
            worksheet.column_dimensions[column_letter].width = 20
